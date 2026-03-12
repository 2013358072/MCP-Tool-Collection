from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from fastmcp import FastMCP

from mcp_toolkit.core import config as _cfg
from mcp_toolkit.providers.base import BaseProvider

# ======================================================================
# 本模块仅处理 .xlsx 格式表格（openpyxl）
# ======================================================================


def _mkparent(p: Path) -> None:
    p.parent.mkdir(parents=True, exist_ok=True)


def _resolve_sheet_path(path: str) -> Path:
    """将表格路径限制在沙箱目录内。"""
    sandbox_root = Path(_cfg.FILESYSTEM_ROOT or Path.cwd()).resolve()
    p = Path(path)
    resolved = (sandbox_root / p).resolve() if not p.is_absolute() else p.resolve()
    try:
        resolved.relative_to(sandbox_root)
    except ValueError as e:
        raise PermissionError(
            f"路径越出沙箱: {resolved}（FILESYSTEM_ROOT={sandbox_root}）"
        ) from e
    return resolved


def _check_xlsx(p: Path) -> Optional[Dict[str, Any]]:
    """校验文件存在且后缀为 .xlsx，不满足时返回错误字典。"""
    if not p.exists() or not p.is_file():
        return {"ok": False, "error": "NOT_FOUND", "path": str(p)}
    if p.suffix.lower() != ".xlsx":
        return {"ok": False, "error": "NOT_XLSX", "detail": f"仅支持 .xlsx 格式，收到: {p.suffix}"}
    return None


def _parse_a1(notation: str) -> Tuple[int, int, int, int]:
    """解析 A1 notation，返回 (min_row, min_col, max_row, max_col)，均为 1-based。"""
    parts = notation.upper().strip().split(":")
    r1, c1 = coordinate_to_tuple(parts[0])
    if len(parts) == 1:
        return r1, c1, r1, c1
    r2, c2 = coordinate_to_tuple(parts[1])
    return min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2)


def _get_sheet(wb: Workbook, sheet: Optional[str]):
    """按名称获取工作表，未指定时返回活动表。"""
    if sheet and sheet in wb.sheetnames:
        return wb[sheet]
    return wb.active


# ---------------------------------------------------------------------- #
# sheets_read_range
# ---------------------------------------------------------------------- #

def _sheets_read_range(
    path: str,
    range_notation: str,
    sheet: Optional[str] = None,
) -> Dict[str, Any]:
    try:
        p = _resolve_sheet_path(path)
    except PermissionError as e:
        return {"ok": False, "error": "PATH_OUTSIDE_SANDBOX", "detail": str(e)}
    if err := _check_xlsx(p):
        return err
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = _get_sheet(wb, sheet)
        min_row, min_col, max_row, max_col = _parse_a1(range_notation)
        rows: List[List[Any]] = []
        for row in ws.iter_rows(
            min_row=min_row, max_row=max_row,
            min_col=min_col, max_col=max_col,
            values_only=True,
        ):
            rows.append(list(row))
        wb.close()
        return {
            "ok": True,
            "path": str(p),
            "sheet": ws.title,
            "range": range_notation,
            "row_count": len(rows),
            "col_count": len(rows[0]) if rows else 0,
            "rows": rows,
        }
    except Exception as e:
        return {"ok": False, "error": "READ_FAILED", "path": str(p), "detail": str(e)}


# ---------------------------------------------------------------------- #
# sheets_write_range
# ---------------------------------------------------------------------- #

def _sheets_write_range(
    path: str,
    range_notation: str,
    rows: List[List[Any]],
    sheet: Optional[str] = None,
) -> Dict[str, Any]:
    """从指定起始单元格开始写入二维数组；文件不存在时自动创建。"""
    try:
        p = _resolve_sheet_path(path)
    except PermissionError as e:
        return {"ok": False, "error": "PATH_OUTSIDE_SANDBOX", "detail": str(e)}
    if p.suffix.lower() != ".xlsx":
        return {"ok": False, "error": "NOT_XLSX", "detail": f"仅支持 .xlsx 格式，收到: {p.suffix}"}
    try:
        _mkparent(p)
        wb = load_workbook(p) if p.exists() else Workbook()
        ws = _get_sheet(wb, sheet)
        min_row, min_col, _, _ = _parse_a1(range_notation)
        for r_idx, row in enumerate(rows):
            for c_idx, value in enumerate(row):
                ws.cell(row=min_row + r_idx, column=min_col + c_idx, value=value)
        wb.save(p)
        return {
            "ok": True,
            "path": str(p),
            "sheet": ws.title,
            "range": range_notation,
            "rows_written": len(rows),
        }
    except Exception as e:
        return {"ok": False, "error": "WRITE_FAILED", "path": str(p), "detail": str(e)}


# ---------------------------------------------------------------------- #
# sheets_append_rows
# ---------------------------------------------------------------------- #

def _sheets_append_rows(
    path: str,
    rows: List[List[Any]],
    sheet: Optional[str] = None,
) -> Dict[str, Any]:
    """追加多行到工作表末尾；文件不存在时自动创建。"""
    try:
        p = _resolve_sheet_path(path)
    except PermissionError as e:
        return {"ok": False, "error": "PATH_OUTSIDE_SANDBOX", "detail": str(e)}
    if p.suffix.lower() != ".xlsx":
        return {"ok": False, "error": "NOT_XLSX", "detail": f"仅支持 .xlsx 格式，收到: {p.suffix}"}
    try:
        _mkparent(p)
        wb = load_workbook(p) if p.exists() else Workbook()
        ws = _get_sheet(wb, sheet)
        for row in rows:
            ws.append(row)
        wb.save(p)
        return {
            "ok": True,
            "path": str(p),
            "sheet": ws.title,
            "rows_appended": len(rows),
            "total_rows": ws.max_row,
        }
    except Exception as e:
        return {"ok": False, "error": "APPEND_FAILED", "path": str(p), "detail": str(e)}


# ---------------------------------------------------------------------- #
# sheets_sort_range
# ---------------------------------------------------------------------- #

def _sheets_sort_range(
    path: str,
    range_notation: str,
    sort_keys: List[Dict[str, Any]],
    sheet: Optional[str] = None,
    has_header: bool = True,
) -> Dict[str, Any]:
    """对指定范围按多字段排序后原地写回。

    sort_keys 每项格式：
        {"col": "A"}                         # 按 A 列升序
        {"col": "B", "ascending": False}     # 按 B 列降序
        {"col": 2}                           # col 也可传 1-based 整数列号
    """
    try:
        p = _resolve_sheet_path(path)
    except PermissionError as e:
        return {"ok": False, "error": "PATH_OUTSIDE_SANDBOX", "detail": str(e)}
    if err := _check_xlsx(p):
        return err
    try:
        wb = load_workbook(p)
        ws = _get_sheet(wb, sheet)
        min_row, min_col, max_row, max_col = _parse_a1(range_notation)

        # 读取全部数据行
        all_rows: List[List[Any]] = []
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
            all_rows.append(list(row))

        if not all_rows:
            return {"ok": True, "path": str(p), "sorted_rows": 0}

        header = all_rows[0] if has_header else None
        data = all_rows[1:] if has_header else all_rows

        # 构建排序 key（支持列字母或 1-based 整数，相对于 min_col）
        def _col_offset(col_spec: Any) -> int:
            if isinstance(col_spec, int):
                return col_spec - min_col          # 传入绝对列号
            col_letter = str(col_spec).upper()
            return column_index_from_string(col_letter) - min_col

        # 多字段稳定排序（从最次要字段开始逐级排）
        for key in reversed(sort_keys):
            offset = _col_offset(key.get("col", 1))
            ascending = bool(key.get("ascending", True))
            data.sort(
                key=lambda r, o=offset: (r[o] is None, r[o] if r[o] is not None else ""),
                reverse=not ascending,
            )

        # 写回（含表头）
        write_rows = ([header] + data) if has_header else data
        for r_idx, row in enumerate(write_rows):
            for c_idx, value in enumerate(row):
                ws.cell(row=min_row + r_idx, column=min_col + c_idx, value=value)

        wb.save(p)
        return {
            "ok": True,
            "path": str(p),
            "sheet": ws.title,
            "range": range_notation,
            "sorted_rows": len(data),
        }
    except Exception as e:
        return {"ok": False, "error": "SORT_FAILED", "path": str(p), "detail": str(e)}


# ---------------------------------------------------------------------- #
# sheets_export_xlsx
# ---------------------------------------------------------------------- #

def _sheets_export_xlsx(
    path: str,
    rows: List[List[Any]],
    headers: Optional[List[str]] = None,
    sheet_name: str = "Sheet1",
    column_widths: Optional[Dict[str, float]] = None,
) -> Dict[str, Any]:
    """根据结构化数据构建并导出 xlsx 文件。

    rows:         二维数组，每项为一行数据
    headers:      列标题（可选），写入第一行并加粗
    sheet_name:   工作表名称
    column_widths 列宽配置，格式 {"A": 20, "B": 15}（可选）
    """
    try:
        p = _resolve_sheet_path(path)
    except PermissionError as e:
        return {"ok": False, "error": "PATH_OUTSIDE_SANDBOX", "detail": str(e)}
    if p.suffix.lower() != ".xlsx":
        return {"ok": False, "error": "NOT_XLSX", "detail": f"输出路径必须以 .xlsx 结尾，收到: {p.suffix}"}
    try:
        _mkparent(p)
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        header_font = Font(bold=True)
        current_row = 1

        if headers:
            for c_idx, title in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=c_idx, value=title)
                cell.font = header_font
            current_row += 1

        for row in rows:
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=current_row, column=c_idx, value=value)
            current_row += 1

        if column_widths:
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter.upper()].width = width

        wb.save(p)
        return {
            "ok": True,
            "path": str(p),
            "sheet": ws.title,
            "header_count": len(headers) if headers else 0,
            "rows_written": len(rows),
            "size_bytes": p.stat().st_size,
        }
    except Exception as e:
        return {"ok": False, "error": "EXPORT_FAILED", "path": str(p), "detail": str(e)}


# ======================================================================
# Provider
# ======================================================================

class SheetsProvider(BaseProvider):
    """Excel 表格操作工具集 Provider（仅处理 .xlsx 格式，依赖 openpyxl）。"""

    @property
    def name(self) -> str:
        return "sheets"

    def is_available(self) -> bool:
        return True

    async def initialize(self) -> None:
        self.logger.local("info", "SheetsProvider 初始化完成")

    def register(self, mcp: FastMCP) -> None:

        @mcp.tool()
        async def sheets_read_range(
            path: str,
            range_notation: str,
            sheet: Optional[str] = None,
        ) -> Dict[str, Any]:
            """读取 xlsx 表格指定范围的数据（A1 notation，如 A1:C10）"""
            return _sheets_read_range(path, range_notation, sheet=sheet)

        @mcp.tool()
        async def sheets_write_range(
            path: str,
            range_notation: str,
            rows: List[List[Any]],
            sheet: Optional[str] = None,
        ) -> Dict[str, Any]:
            """从指定起始单元格写入二维数组；文件不存在时自动创建"""
            return _sheets_write_range(path, range_notation, rows, sheet=sheet)

        @mcp.tool()
        async def sheets_append_rows(
            path: str,
            rows: List[List[Any]],
            sheet: Optional[str] = None,
        ) -> Dict[str, Any]:
            """追加多行到 xlsx 工作表末尾；文件不存在时自动创建"""
            return _sheets_append_rows(path, rows, sheet=sheet)

        @mcp.tool()
        async def sheets_sort_range(
            path: str,
            range_notation: str,
            sort_keys: List[Dict[str, Any]],
            sheet: Optional[str] = None,
            has_header: bool = True,
        ) -> Dict[str, Any]:
            """对指定范围按多字段排序后原地写回。
            sort_keys 每项: {"col": "A", "ascending": true} 或 {"col": 2, "ascending": false}
            has_header: 首行是否为标题（排序时跳过）
            """
            return _sheets_sort_range(path, range_notation, sort_keys, sheet=sheet, has_header=has_header)

        @mcp.tool()
        async def sheets_export_xlsx(
            path: str,
            rows: List[List[Any]],
            headers: Optional[List[str]] = None,
            sheet_name: str = "Sheet1",
            column_widths: Optional[Dict[str, float]] = None,
        ) -> Dict[str, Any]:
            """根据结构化数据构建并导出 xlsx 文件。
            headers: 列标题列表，写入第一行并加粗（可选）
            column_widths: 列宽配置，如 {"A": 20, "B": 15}（可选）
            """
            return _sheets_export_xlsx(path, rows, headers=headers, sheet_name=sheet_name, column_widths=column_widths)
